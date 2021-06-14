from openpyxl import load_workbook
from openpyxl.cell import cell

SCRIPT = 0
NUMBER_MAPPING = 1
CONVERSATION_DATABASE = 2
GAME = 3


def get_all_choices(number_mapping):

    INDEX = 1
    ROW_INDEX = 2
    OPTION_A = 2
    OPTION_B = 3

    choices = []
    while number_mapping.cell(row=ROW_INDEX,column=INDEX).value != None:
        cid = number_mapping.cell(row=ROW_INDEX, column=INDEX).value
        optionA = number_mapping.cell(row=ROW_INDEX, column=OPTION_A).value
        optionB = number_mapping.cell(row=ROW_INDEX, column=OPTION_B).value
        choices.append((cid, optionA))
        choices.append((cid, optionB))
        ROW_INDEX += 1
    return choices

def get_the_rest(script):

    PRIMARY_TAG1 = 11 
    PRIMARY_TAG2 = 12

    SECONDARY_TAG1 = 13
    SECONDARY_TAG2 = 14
    SECONDARY_TAG3 = 15

    ROW_INDEX = 1
    INDEX = 1
    TEXT_COL = 2

    conversation_text = {}
    edge_text = {}
    edge_tags = {}
    while script.cell(row=ROW_INDEX,column=INDEX).value != "###":
        cell_text = str(script.cell(row=ROW_INDEX, column=INDEX).value)
        if cell_text is None:
            ROW_INDEX += 1
            continue
        else:
            text = script.cell(row=ROW_INDEX, column=TEXT_COL).value
            if text is None: 
                ROW_INDEX +=1
                continue
            if cell_text.startswith("If") or cell_text.startswith("When"):
                try:
                    cid = int(cell_text.split(" ")[1])
                    conversation_text[cid] = text
                except:
                    print(cell_text)
            elif cell_text.isnumeric():
                cid=int(cell_text)
                edge_text[cid] = text
                edge_tags[cid] = { "p1": script.cell(row=ROW_INDEX, column=PRIMARY_TAG1).value,
                                   "p2": script.cell(row=ROW_INDEX, column=PRIMARY_TAG2).value,
                                   "s1": script.cell(row=ROW_INDEX, column=SECONDARY_TAG1).value,
                                   "s2": script.cell(row=ROW_INDEX, column=SECONDARY_TAG2).value,
                                   "s3": script.cell(row=ROW_INDEX, column=SECONDARY_TAG3).value}
            else:
                pass
        ROW_INDEX += 1

    return edge_text, conversation_text, edge_tags

def parse(path_to_excel):
    wb = load_workbook(path_to_excel)

    script = wb.worksheets[SCRIPT]
    number_mapping = wb.worksheets[NUMBER_MAPPING]

    choices = get_all_choices(number_mapping)
    edge_text, conversation_text, edge_tags = get_the_rest(script)

    return choices, edge_text, conversation_text, edge_tags


                



        

